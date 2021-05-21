import openpyxl


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return (ws.max_row)

def getColumnCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return (ws.max_column)

def readData(file, sheetName, rownum, colnum):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.cell(row=rownum, column=colnum).value

def writeData(file, sheetName, rownum, colnum, data):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(row=rownum, column=colnum).value = data
    wb.save(file)